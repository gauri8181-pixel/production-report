from __future__ import annotations

import io
import re
from copy import copy
from datetime import datetime, date, timedelta
from collections import defaultdict

import altair as alt
import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList


# =========================
# 기본 설정
# =========================
EXCLUDE_PROD_LINES = {
    "(소파) 고객만족",
    "(소파) 재단",
    "(의자) 재단",
    "",
    None,
}

FURSYS_INHOUSE_SUPPLIER = "시디즈안성"
ALLOSO_INHOUSE_SUPPLIER_EMPTY_OK = True

BRANDS = ["알로소", "일룸", "퍼시스"]
MAKERS = ["내작", "외주"]

UNIFORM_END_COL = 18
CHART_WIDTH = 16
CHART_HEIGHT = 8.5
CHART_ROW_GAP = 19
CHART_COL_LEFT = "A"
CHART_COL_RIGHT = "J"


# =========================
# 스타일
# =========================
FONT_9 = Font(name="맑은 고딕", size=9)
BOLD_9 = Font(name="맑은 고딕", size=9, bold=True)
TITLE_11 = Font(name="맑은 고딕", size=11, bold=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN = Side(style="thin")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_HEADER = PatternFill("solid", fgColor="D9EAF7")
FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")


# =========================
# 공통 유틸
# =========================
def norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).strip().lower()


def try_parse_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip()
    cand = s.replace(".", "-").replace("/", "-")

    if len(cand) >= 10:
        try:
            return datetime.strptime(cand[:10], "%Y-%m-%d").date()
        except Exception:
            pass

    if len(s) == 8 and s.isdigit():
        try:
            return datetime.strptime(s, "%Y%m%d").date()
        except Exception:
            pass

    return None


def to_number(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except Exception:
        return 0.0


def month_key(d: date):
    return d.strftime("%Y-%m")


def brand_from_itemcode(code: str) -> str:
    c = (code or "").strip()
    if not c:
        return "기타"
    ch = c[0].lower()
    if ch == "a":
        return "알로소"
    if ch in ("h", "i", "d"):
        return "일룸"
    if ch in ("c", "z"):
        return "퍼시스"
    return "기타"


def daterange(d1: date, d2: date):
    d = d1
    while d <= d2:
        yield d
        d += timedelta(days=1)


def get_month_keys(start_date: date, end_date: date):
    months = []
    y, m = start_date.year, start_date.month
    while (y < end_date.year) or (y == end_date.year and m <= end_date.month):
        months.append(f"{y:04d}-{m:02d}")
        m += 1
        if m == 13:
            m = 1
            y += 1
    return months


def get_week_ranges(start_date: date, end_date: date):
    start_monday = start_date - timedelta(days=start_date.weekday())
    end_sunday = end_date + timedelta(days=(6 - end_date.weekday()))
    weeks = []
    cur = start_monday
    while cur <= end_sunday:
        weeks.append((cur, cur + timedelta(days=6)))
        cur += timedelta(days=7)
    return weeks


def autosize_columns(ws, min_w=8, max_w=20):
    for col in range(1, ws.max_column + 1):
        mx = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row, col).value
            if val is None:
                continue
            mx = max(mx, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = max(min_w, min(max_w, mx + 2))


def map_headers(ws):
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        headers[norm(v)] = c
    return headers


def read_rows(ws):
    headers = map_headers(ws)
    rows = []
    for r in range(2, ws.max_row + 1):
        row_obj = {}
        empty = True
        for key_norm, col in headers.items():
            v = ws.cell(r, col).value
            if v is not None and v != "":
                empty = False
            row_obj[key_norm] = v
        if not empty:
            rows.append(row_obj)
    return headers, rows


def safe_get(row, header_candidates):
    for cand in header_candidates:
        key = norm(cand)
        if key in row:
            return row[key]
    return None


def set_cell_style(cell, font=None, fill=None, align=None, border=None, numfmt=None):
    if font is not None:
        cell.font = copy(font)
    if fill is not None:
        cell.fill = copy(fill)
    if align is not None:
        cell.alignment = copy(align)
    if border is not None:
        cell.border = copy(border)
    if numfmt is not None:
        cell.number_format = numfmt


def apply_table_style(ws, r1, c1, r2, c2, header_rows=1, total_row=None):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            set_cell_style(ws.cell(r, c), font=FONT_9, align=CENTER, border=BORDER_THIN)

    for rr in range(r1, r1 + header_rows):
        for c in range(c1, c2 + 1):
            set_cell_style(ws.cell(rr, c), font=BOLD_9, fill=FILL_HEADER, align=CENTER, border=BORDER_THIN)

    if total_row is not None:
        for c in range(c1, c2 + 1):
            set_cell_style(ws.cell(total_row, c), font=BOLD_9, fill=FILL_TOTAL, align=CENTER, border=BORDER_THIN)


def format_number_cells(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"


def pad_table_right(ws, r1, r2, current_end_col, target_end_col):
    if current_end_col >= target_end_col:
        return
    for r in range(r1, r2 + 1):
        for c in range(current_end_col + 1, target_end_col + 1):
            set_cell_style(ws.cell(r, c), font=FONT_9, align=CENTER, border=BORDER_THIN)


def write_title(ws, row, col, text):
    cell = ws.cell(row, col)
    cell.value = text
    cell.font = copy(TITLE_11)
    cell.alignment = copy(LEFT)
    return row + 1


# =========================
# 웹 표시용 보조 함수
# =========================
def add_total_row(df: pd.DataFrame, label_col: str, label_name: str = "합계") -> pd.DataFrame:
    if df.empty:
        return df

    total_data = {}
    for col in df.columns:
        if col == label_col:
            total_data[col] = label_name
        else:
            if pd.api.types.is_numeric_dtype(df[col]):
                total_data[col] = df[col].sum()
            else:
                total_data[col] = ""

    total_df = pd.DataFrame([total_data])
    return pd.concat([df, total_df], ignore_index=True)


def format_df_for_display(df: pd.DataFrame):
    if df.empty:
        return df.style

    num_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    fmt = {c: "{:,.0f}" for c in num_cols}
    return df.style.format(fmt)


def filter_order_rows_by_selection(df: pd.DataFrame, label_col: str, selected_brands: list[str], selected_makers: list[str]) -> pd.DataFrame:
    if df.empty or label_col not in df.columns:
        return df

    def keep_row(v):
        s = str(v)
        if "-" not in s:
            return True
        brand, maker = s.split("-", 1)
        return (brand in selected_brands) and (maker in selected_makers)

    return df[df[label_col].apply(keep_row)].reset_index(drop=True)


# =========================
# 데이터 집계 - 수주
# =========================
def collect_orders(wb):
    agg_daily = defaultdict(lambda: {"qty": 0.0, "amt": 0.0})

    DATE_COLS = ["주문일자", "주문일", "수주일", "오더일자"]
    QTY_COLS = ["수주량", "수주수량", "주문수량", "수량"]
    AMT_COLS = ["입고금액", "금액", "수주금액"]
    BRAND_COLS = ["브랜드"]
    SUP_COLS = ["공급처", "협력사", "업체", "공급업체"]

    for ws_name in wb.sheetnames:
        if "수주" not in ws_name:
            continue

        ws = wb[ws_name]
        headers, rows = read_rows(ws)

        sheet_brand = None
        if "알로소" in ws_name:
            sheet_brand = "알로소"
        elif "일룸" in ws_name:
            sheet_brand = "일룸"
        elif "퍼시스" in ws_name:
            sheet_brand = "퍼시스"

        sheet_maker = None
        if "내작" in ws_name:
            sheet_maker = "내작"
        elif "외주" in ws_name or "상품" in ws_name:
            sheet_maker = "외주"

        if not any(norm(x) in headers for x in map(norm, DATE_COLS)):
            continue
        if not any(norm(x) in headers for x in map(norm, QTY_COLS)):
            continue
        if not any(norm(x) in headers for x in map(norm, AMT_COLS)):
            continue

        for row in rows:
            dt = try_parse_date(safe_get(row, DATE_COLS))
            if dt is None:
                continue

            qty = to_number(safe_get(row, QTY_COLS))
            amt = to_number(safe_get(row, AMT_COLS))

            b = safe_get(row, BRAND_COLS)
            b = (str(b).strip() if b is not None else "")
            if b == "":
                b = sheet_brand if sheet_brand else "기타"

            supplier = safe_get(row, SUP_COLS)
            supplier = (str(supplier).strip() if supplier is not None else "")

            maker = sheet_maker
            if maker is None:
                if b == "퍼시스":
                    maker = "내작" if supplier == FURSYS_INHOUSE_SUPPLIER else "외주"
                elif b == "알로소" and ALLOSO_INHOUSE_SUPPLIER_EMPTY_OK and supplier == "":
                    maker = "내작"
                else:
                    maker = "내작" if supplier == FURSYS_INHOUSE_SUPPLIER else "외주"

            agg_daily[(b, maker, dt)]["qty"] += qty
            agg_daily[(b, maker, dt)]["amt"] += amt

    return agg_daily


# =========================
# 데이터 집계 - 생산
# =========================
def collect_production_plan_actual(wb):
    agg = defaultdict(lambda: {"plan_qty": 0.0, "plan_amt": 0.0, "act_qty": 0.0, "act_amt": 0.0})

    C_ITEM = ["품목코드", "제품코드", "단품코드"]
    C_LINE = ["생산라인"]
    C_PLANQ = ["계획량", "계획수량"]
    C_ACTQ = ["생산량", "실적수량"]
    C_PLAN_DATE = ["최초포장계획일", "최초 포장 계획일", "최초포장계획일자"]
    C_ACT_DATE = ["포장계획일", "포장 계획일", "포장계획일자"]
    C_UNIT = ["입고단가", "단가"]

    for ws_name in wb.sheetnames:
        if "생산" not in ws_name:
            continue

        ws = wb[ws_name]
        headers, rows = read_rows(ws)

        needed_groups = [C_ITEM, C_LINE, C_PLANQ, C_ACTQ, C_PLAN_DATE, C_ACT_DATE, C_UNIT]
        if any(not any(norm(x) in headers for x in map(norm, grp)) for grp in needed_groups):
            continue

        for row in rows:
            all_text = " ".join([str(v) for v in row.values() if v is not None])
            if "sub" in all_text.lower() and "total" in all_text.lower():
                continue

            line = safe_get(row, C_LINE)
            line = str(line).strip() if line is not None else ""
            if line in EXCLUDE_PROD_LINES:
                continue

            item = safe_get(row, C_ITEM)
            item = str(item).strip() if item is not None else ""
            if not item:
                continue

            brand = brand_from_itemcode(item)

            unit = to_number(safe_get(row, C_UNIT))
            plan_qty = to_number(safe_get(row, C_PLANQ))
            act_qty = to_number(safe_get(row, C_ACTQ))

            plan_date = try_parse_date(safe_get(row, C_PLAN_DATE))
            act_date = try_parse_date(safe_get(row, C_ACT_DATE))

            if plan_date:
                agg[(brand, plan_date)]["plan_qty"] += plan_qty
                agg[(brand, plan_date)]["plan_amt"] += plan_qty * unit

            if act_date:
                agg[(brand, act_date)]["act_qty"] += act_qty
                agg[(brand, act_date)]["act_amt"] += act_qty * unit

    return agg


# =========================
# 엑셀 차트
# =========================
def make_excel_style_combo_chart(ws, title, x_title, y_left, y_right, table_top_row, table_left_col, n_rows):
    min_row = table_top_row
    max_row = table_top_row + n_rows - 1
    min_col = table_left_col

    cat_ref = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    qty_ref = Reference(ws, min_col=min_col + 1, min_row=min_row, max_col=min_col + 2, max_row=max_row)
    amt_ref = Reference(ws, min_col=min_col + 3, min_row=min_row, max_col=min_col + 4, max_row=max_row)

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = title
    bar.style = 10
    bar.y_axis.title = y_left
    bar.x_axis.title = x_title
    bar.gapWidth = 70
    bar.overlap = 0
    bar.legend.position = "b"
    bar.width = CHART_WIDTH
    bar.height = CHART_HEIGHT
    bar.add_data(qty_ref, titles_from_data=True)
    bar.set_categories(cat_ref)

    line = LineChart()
    line.add_data(amt_ref, titles_from_data=True)
    line.set_categories(cat_ref)
    line.y_axis.axId = 200
    line.y_axis.title = y_right
    line.y_axis.crosses = "max"

    try:
        for s in line.series:
            s.marker.symbol = "circle"
            s.marker.size = 6
            s.graphicalProperties.line.width = 19050
    except Exception:
        pass

    bar.dLbls = DataLabelList()
    bar.dLbls.showVal = False
    bar += line
    return bar


# =========================
# 웹 그래프
# =========================
def build_chart_data(
    orders_agg,
    prod_agg,
    month_start,
    month_end,
    week_start,
    week_end,
    day_start,
    day_end,
    prod_summary_start,
    prod_summary_end,
    prod_detail_start,
    prod_detail_end,
):
    months = get_month_keys(month_start, month_end)
    monthly_rows = []
    for mk in months:
        in_qty = out_qty = in_amt = out_amt = 0.0
        for (bb, mm, dt), vals in orders_agg.items():
            if month_key(dt) != mk:
                continue
            if mm == "내작":
                in_qty += vals["qty"]
                in_amt += vals["amt"]
            else:
                out_qty += vals["qty"]
                out_amt += vals["amt"]

        monthly_rows.append({
            "기간": mk,
            "내작수량": in_qty,
            "외주수량": out_qty,
            "내작금액": in_amt,
            "외주금액": out_amt,
        })
    df_monthly = pd.DataFrame(monthly_rows)

    weeks = get_week_ranges(week_start, week_end)
    if len(weeks) > 5:
        weeks = weeks[-5:]

    weekly_rows = []
    for w_s, w_e in weeks:
        in_qty = out_qty = in_amt = out_amt = 0.0
        for (bb, mm, dt), vals in orders_agg.items():
            if not (w_s <= dt <= w_e):
                continue
            if mm == "내작":
                in_qty += vals["qty"]
                in_amt += vals["amt"]
            else:
                out_qty += vals["qty"]
                out_amt += vals["amt"]

        weekly_rows.append({
            "기간": f"{w_s.strftime('%m/%d')}~{w_e.strftime('%m/%d')}",
            "내작수량": in_qty,
            "외주수량": out_qty,
            "내작금액": in_amt,
            "외주금액": out_amt,
        })
    df_weekly = pd.DataFrame(weekly_rows)

    daily_rows = []
    for d in daterange(day_start, day_end):
        in_qty = out_qty = in_amt = out_amt = 0.0
        for b in BRANDS:
            v1 = orders_agg.get((b, "내작", d), {"qty": 0.0, "amt": 0.0})
            v2 = orders_agg.get((b, "외주", d), {"qty": 0.0, "amt": 0.0})
            in_qty += v1["qty"]
            in_amt += v1["amt"]
            out_qty += v2["qty"]
            out_amt += v2["amt"]

        daily_rows.append({
            "기간": d.strftime("%m/%d"),
            "내작수량": in_qty,
            "외주수량": out_qty,
            "내작금액": in_amt,
            "외주금액": out_amt,
        })
    df_daily = pd.DataFrame(daily_rows)

    prod_summary_rows = []
    for b in BRANDS:
        plan_q = plan_a = act_q = act_a = 0.0
        for d in daterange(prod_summary_start, prod_summary_end):
            v = prod_agg.get((b, d), {"plan_qty": 0.0, "plan_amt": 0.0, "act_qty": 0.0, "act_amt": 0.0})
            plan_q += v["plan_qty"]
            plan_a += v["plan_amt"]
            act_q += v["act_qty"]
            act_a += v["act_amt"]

        prod_summary_rows.append({
            "브랜드": b,
            "계획수량": plan_q,
            "실적수량": act_q,
            "계획금액": plan_a,
            "실적금액": act_a,
        })
    df_prod_summary = pd.DataFrame(prod_summary_rows)

    prod_detail_rows = []
    for d in daterange(prod_detail_start, prod_detail_end):
        plan_q = plan_a = act_q = act_a = 0.0
        for b in BRANDS:
            v = prod_agg.get((b, d), {"plan_qty": 0.0, "plan_amt": 0.0, "act_qty": 0.0, "act_amt": 0.0})
            plan_q += v["plan_qty"]
            plan_a += v["plan_amt"]
            act_q += v["act_qty"]
            act_a += v["act_amt"]

        prod_detail_rows.append({
            "기간": d.strftime("%m/%d"),
            "계획수량": plan_q,
            "실적수량": act_q,
            "계획금액": plan_a,
            "실적금액": act_a,
        })
    df_prod_detail = pd.DataFrame(prod_detail_rows)

    return df_monthly, df_weekly, df_daily, df_prod_summary, df_prod_detail


def show_combo_chart(df, title, x_col, qty_cols, amt_cols):
    if df.empty:
        st.info("표시할 데이터가 없습니다.")
        return

    qty_long = df[[x_col] + qty_cols].melt(id_vars=[x_col], var_name="구분", value_name="값")
    amt_long = df[[x_col] + amt_cols].melt(id_vars=[x_col], var_name="구분", value_name="값")

    bar = alt.Chart(qty_long).mark_bar().encode(
        x=alt.X(f"{x_col}:N", title=x_col),
        y=alt.Y("값:Q", title="수량"),
        color=alt.Color("구분:N", title="수량 구분"),
        xOffset="구분:N",
        tooltip=[x_col, "구분", alt.Tooltip("값:Q", format=",.0f")]
    )

    line = alt.Chart(amt_long).mark_line(point=True).encode(
        x=alt.X(f"{x_col}:N", title=x_col),
        y=alt.Y("값:Q", title="금액"),
        color=alt.Color("구분:N", title="금액 구분"),
        tooltip=[x_col, "구분", alt.Tooltip("값:Q", format=",.0f")]
    )

    chart = alt.layer(bar, line).resolve_scale(
        y="independent"
    ).properties(
        title=title,
        height=380
    )

    st.altair_chart(chart, use_container_width=True)


# =========================
# 엑셀 시트 작성
# =========================
def write_orders_section(ws, top_row, orders_agg, month_start, month_end, week_start, week_end, day_start, day_end):
    charts = []
    r = top_row

    months = get_month_keys(month_start, month_end)
    weeks = get_week_ranges(week_start, week_end)
    if len(weeks) > 5:
        weeks = weeks[-5:]
    days = list(daterange(day_start, day_end))

    r = write_title(ws, r, 1, "1) 월별 수주현황")
    top = r
    ws.cell(top, 1).value = "브랜드-구분"
    c = 2
    for m in months:
        ws.cell(top, c).value = m
        ws.merge_cells(start_row=top, start_column=c, end_row=top, end_column=c + 1)
        c += 2
    ws.cell(top, c).value = "합계"
    ws.merge_cells(start_row=top, start_column=c, end_row=top, end_column=c + 1)

    top2 = top + 1
    ws.cell(top2, 1).value = ""
    c = 2
    for _ in months:
        ws.cell(top2, c).value = "수량"
        ws.cell(top2, c + 1).value = "금액"
        c += 2
    ws.cell(top2, c).value = "수량"
    ws.cell(top2, c + 1).value = "금액"

    rr = top + 2
    for b in BRANDS:
        for mkr in MAKERS:
            ws.cell(rr, 1).value = f"{b}-{mkr}"
            c = 2
            total_qty = total_amt = 0.0
            for mk in months:
                qty = amt = 0.0
                for (bb, mm, dt), vals in orders_agg.items():
                    if bb == b and mm == mkr and month_key(dt) == mk:
                        qty += vals["qty"]
                        amt += vals["amt"]
                ws.cell(rr, c).value = qty
                ws.cell(rr, c + 1).value = amt
                total_qty += qty
                total_amt += amt
                c += 2
            ws.cell(rr, c).value = total_qty
            ws.cell(rr, c + 1).value = total_amt
            rr += 1

    ws.cell(rr, 1).value = "합계"
    c = 2
    for _ in months:
        ws.cell(rr, c).value = f"=SUM({get_column_letter(c)}{top+2}:{get_column_letter(c)}{rr-1})"
        ws.cell(rr, c + 1).value = f"=SUM({get_column_letter(c+1)}{top+2}:{get_column_letter(c+1)}{rr-1})"
        c += 2
    ws.cell(rr, c).value = f"=SUM({get_column_letter(c)}{top+2}:{get_column_letter(c)}{rr-1})"
    ws.cell(rr, c + 1).value = f"=SUM({get_column_letter(c+1)}{top+2}:{get_column_letter(c+1)}{rr-1})"

    apply_table_style(ws, top, 1, rr, c + 1, header_rows=2, total_row=rr)
    format_number_cells(ws, top + 2, 2, rr, c + 1)
    pad_table_right(ws, top, rr, c + 1, UNIFORM_END_COL)
    r = rr + 2

    chart_top = r
    ws.cell(chart_top, 1).value = "월"
    ws.cell(chart_top, 2).value = "내작수량"
    ws.cell(chart_top, 3).value = "외주수량"
    ws.cell(chart_top, 4).value = "내작금액"
    ws.cell(chart_top, 5).value = "외주금액"
    rr = chart_top + 1
    for mk in months:
        in_qty = out_qty = in_amt = out_amt = 0.0
        for (bb, mm, dt), vals in orders_agg.items():
            if month_key(dt) != mk:
                continue
            if mm == "내작":
                in_qty += vals["qty"]
                in_amt += vals["amt"]
            else:
                out_qty += vals["qty"]
                out_amt += vals["amt"]
        ws.cell(rr, 1).value = mk
        ws.cell(rr, 2).value = in_qty
        ws.cell(rr, 3).value = out_qty
        ws.cell(rr, 4).value = in_amt
        ws.cell(rr, 5).value = out_amt
        rr += 1
    apply_table_style(ws, chart_top, 1, rr - 1, 5, header_rows=1)
    format_number_cells(ws, chart_top + 1, 2, rr - 1, 5)
    charts.append(("월별 수주 추세", chart_top, rr - chart_top))
    r = rr + 2

    r = write_title(ws, r, 1, "2) 최근 5주간 주간 합계")
    top = r
    ws.cell(top, 1).value = "브랜드-구분"
    c = 2
    for ws_date, we_date in weeks:
        title = f"{ws_date.strftime('%m/%d')}~{we_date.strftime('%m/%d')}"
        ws.cell(top, c).value = title
        ws.merge_cells(start_row=top, start_column=c, end_row=top, end_column=c + 1)
        c += 2
    ws.cell(top, c).value = "합계"
    ws.merge_cells(start_row=top, start_column=c, end_row=top, end_column=c + 1)

    top2 = top + 1
    ws.cell(top2, 1).value = ""
    c = 2
    for _ in weeks:
        ws.cell(top2, c).value = "수량"
        ws.cell(top2, c + 1).value = "금액"
        c += 2
    ws.cell(top2, c).value = "수량"
    ws.cell(top2, c + 1).value = "금액"

    rr = top + 2
    for b in BRANDS:
        for mkr in MAKERS:
            ws.cell(rr, 1).value = f"{b}-{mkr}"
            c = 2
            total_qty = total_amt = 0.0
            for w_s, w_e in weeks:
                qty = amt = 0.0
                for d in daterange(w_s, w_e):
                    v = orders_agg.get((b, mkr, d), {"qty": 0.0, "amt": 0.0})
                    qty += v["qty"]
                    amt += v["amt"]
                ws.cell(rr, c).value = qty
                ws.cell(rr, c + 1).value = amt
                total_qty += qty
                total_amt += amt
                c += 2
            ws.cell(rr, c).value = total_qty
            ws.cell(rr, c + 1).value = total_amt
            rr += 1

    ws.cell(rr, 1).value = "합계"
    c = 2
    for _ in weeks:
        ws.cell(rr, c).value = f"=SUM({get_column_letter(c)}{top+2}:{get_column_letter(c)}{rr-1})"
        ws.cell(rr, c + 1).value = f"=SUM({get_column_letter(c+1)}{top+2}:{get_column_letter(c+1)}{rr-1})"
        c += 2
    ws.cell(rr, c).value = f"=SUM({get_column_letter(c)}{top+2}:{get_column_letter(c)}{rr-1})"
    ws.cell(rr, c + 1).value = f"=SUM({get_column_letter(c+1)}{top+2}:{get_column_letter(c+1)}{rr-1})"

    apply_table_style(ws, top, 1, rr, c + 1, header_rows=2, total_row=rr)
    format_number_cells(ws, top + 2, 2, rr, c + 1)
    pad_table_right(ws, top, rr, c + 1, UNIFORM_END_COL)
    r = rr + 2

    chart_top = r
    ws.cell(chart_top, 1).value = "주간"
    ws.cell(chart_top, 2).value = "내작수량"
    ws.cell(chart_top, 3).value = "외주수량"
    ws.cell(chart_top, 4).value = "내작금액"
    ws.cell(chart_top, 5).value = "외주금액"
    rr = chart_top + 1
    for w_s, w_e in weeks:
        in_qty = out_qty = in_amt = out_amt = 0.0
        for (bb, mm, dt), vals in orders_agg.items():
            if not (w_s <= dt <= w_e):
                continue
            if mm == "내작":
                in_qty += vals["qty"]
                in_amt += vals["amt"]
            else:
                out_qty += vals["qty"]
                out_amt += vals["amt"]
        label = f"{w_s.strftime('%m/%d')}~{w_e.strftime('%m/%d')}"
        ws.cell(rr, 1).value = label
        ws.cell(rr, 2).value = in_qty
        ws.cell(rr, 3).value = out_qty
        ws.cell(rr, 4).value = in_amt
        ws.cell(rr, 5).value = out_amt
        rr += 1
    apply_table_style(ws, chart_top, 1, rr - 1, 5, header_rows=1)
    format_number_cells(ws, chart_top + 1, 2, rr - 1, 5)
    charts.append(("5주 주간 합계 추세", chart_top, rr - chart_top))
    r = rr + 2

    r = write_title(ws, r, 1, "3) 지난 1주일 수주 현황")
    top = r
    ws.cell(top, 1).value = "브랜드-구분"
    c = 2
    for d in days:
        ws.cell(top, c).value = d.strftime("%m/%d")
        ws.merge_cells(start_row=top, start_column=c, end_row=top, end_column=c + 1)
        c += 2
    ws.cell(top, c).value = "합계"
    ws.merge_cells(start_row=top, start_column=c, end_row=top, end_column=c + 1)

    top2 = top + 1
    ws.cell(top2, 1).value = ""
    c = 2
    for _ in days:
        ws.cell(top2, c).value = "수량"
        ws.cell(top2, c + 1).value = "금액"
        c += 2
    ws.cell(top2, c).value = "수량"
    ws.cell(top2, c + 1).value = "금액"

    rr = top + 2
    for b in BRANDS:
        for mkr in MAKERS:
            ws.cell(rr, 1).value = f"{b}-{mkr}"
            c = 2
            total_qty = total_amt = 0.0
            for d in days:
                v = orders_agg.get((b, mkr, d), {"qty": 0.0, "amt": 0.0})
                ws.cell(rr, c).value = v["qty"]
                ws.cell(rr, c + 1).value = v["amt"]
                total_qty += v["qty"]
                total_amt += v["amt"]
                c += 2
            ws.cell(rr, c).value = total_qty
            ws.cell(rr, c + 1).value = total_amt
            rr += 1

    ws.cell(rr, 1).value = "합계"
    c = 2
    for _ in days:
        ws.cell(rr, c).value = f"=SUM({get_column_letter(c)}{top+2}:{get_column_letter(c)}{rr-1})"
        ws.cell(rr, c + 1).value = f"=SUM({get_column_letter(c+1)}{top+2}:{get_column_letter(c+1)}{rr-1})"
        c += 2
    ws.cell(rr, c).value = f"=SUM({get_column_letter(c)}{top+2}:{get_column_letter(c)}{rr-1})"
    ws.cell(rr, c + 1).value = f"=SUM({get_column_letter(c+1)}{top+2}:{get_column_letter(c+1)}{rr-1})"

    apply_table_style(ws, top, 1, rr, c + 1, header_rows=2, total_row=rr)
    format_number_cells(ws, top + 2, 2, rr, c + 1)
    pad_table_right(ws, top, rr, c + 1, UNIFORM_END_COL)
    r = rr + 2

    chart_top = r
    ws.cell(chart_top, 1).value = "일자"
    ws.cell(chart_top, 2).value = "내작수량"
    ws.cell(chart_top, 3).value = "외주수량"
    ws.cell(chart_top, 4).value = "내작금액"
    ws.cell(chart_top, 5).value = "외주금액"
    rr = chart_top + 1
    for d in days:
        in_qty = out_qty = in_amt = out_amt = 0.0
        for b in BRANDS:
            v1 = orders_agg.get((b, "내작", d), {"qty": 0.0, "amt": 0.0})
            v2 = orders_agg.get((b, "외주", d), {"qty": 0.0, "amt": 0.0})
            in_qty += v1["qty"]
            in_amt += v1["amt"]
            out_qty += v2["qty"]
            out_amt += v2["amt"]
        ws.cell(rr, 1).value = d.strftime("%m/%d")
        ws.cell(rr, 2).value = in_qty
        ws.cell(rr, 3).value = out_qty
        ws.cell(rr, 4).value = in_amt
        ws.cell(rr, 5).value = out_amt
        rr += 1
    apply_table_style(ws, chart_top, 1, rr - 1, 5, header_rows=1)
    format_number_cells(ws, chart_top + 1, 2, rr - 1, 5)
    charts.append(("지난 1주일 수주 추세", chart_top, rr - chart_top))
    r = rr + 2

    return r, charts


def write_production_section(ws, top_row, prod_agg, prod_summary_start, prod_summary_end, prod_detail_start, prod_detail_end):
    charts = []
    r = top_row

    r = write_title(ws, r, 1, "4) 생산 요약")
    top = r
    ws.cell(top, 1).value = "브랜드"
    ws.cell(top, 2).value = "계획수량"
    ws.cell(top, 3).value = "계획금액"
    ws.cell(top, 4).value = "실적수량"
    ws.cell(top, 5).value = "실적금액"

    rr = top + 1
    for b in BRANDS:
        plan_q = plan_a = act_q = act_a = 0.0
        for d in daterange(prod_summary_start, prod_summary_end):
            v = prod_agg.get((b, d), {"plan_qty": 0.0, "plan_amt": 0.0, "act_qty": 0.0, "act_amt": 0.0})
            plan_q += v["plan_qty"]
            plan_a += v["plan_amt"]
            act_q += v["act_qty"]
            act_a += v["act_amt"]
        ws.cell(rr, 1).value = b
        ws.cell(rr, 2).value = plan_q
        ws.cell(rr, 3).value = plan_a
        ws.cell(rr, 4).value = act_q
        ws.cell(rr, 5).value = act_a
        rr += 1

    ws.cell(rr, 1).value = "합계"
    for c in range(2, 6):
        col = get_column_letter(c)
        ws.cell(rr, c).value = f"=SUM({col}{top+1}:{col}{rr-1})"

    apply_table_style(ws, top, 1, rr, 5, header_rows=1, total_row=rr)
    format_number_cells(ws, top + 1, 2, rr, 5)
    pad_table_right(ws, top, rr, 5, UNIFORM_END_COL)
    r = rr + 2

    chart_top = r
    ws.cell(chart_top, 1).value = "브랜드"
    ws.cell(chart_top, 2).value = "계획수량"
    ws.cell(chart_top, 3).value = "실적수량"
    ws.cell(chart_top, 4).value = "계획금액"
    ws.cell(chart_top, 5).value = "실적금액"
    rr = chart_top + 1
    for b in BRANDS:
        plan_q = plan_a = act_q = act_a = 0.0
        for d in daterange(prod_summary_start, prod_summary_end):
            v = prod_agg.get((b, d), {"plan_qty": 0.0, "plan_amt": 0.0, "act_qty": 0.0, "act_amt": 0.0})
            plan_q += v["plan_qty"]
            plan_a += v["plan_amt"]
            act_q += v["act_qty"]
            act_a += v["act_amt"]
        ws.cell(rr, 1).value = b
        ws.cell(rr, 2).value = plan_q
        ws.cell(rr, 3).value = act_q
        ws.cell(rr, 4).value = plan_a
        ws.cell(rr, 5).value = act_a
        rr += 1
    apply_table_style(ws, chart_top, 1, rr - 1, 5, header_rows=1)
    format_number_cells(ws, chart_top + 1, 2, rr - 1, 5)
    charts.append(("생산 요약 계획/실적", chart_top, rr - chart_top))
    r = rr + 2

    dates = list(daterange(prod_detail_start, prod_detail_end))
    r = write_title(ws, r, 1, "5) 생산 상세")
    top = r
    ws.cell(top, 1).value = "브랜드"
    c = 2
    for d in dates:
        ws.cell(top, c).value = d.strftime("%m/%d")
        ws.merge_cells(start_row=top, start_column=c, end_row=top, end_column=c + 3)
        c += 4
    ws.cell(top, c).value = "합계"
    ws.merge_cells(start_row=top, start_column=c, end_row=top, end_column=c + 3)

    top2 = top + 1
    ws.cell(top2, 1).value = ""
    c = 2
    for _ in dates:
        ws.cell(top2, c + 0).value = "계획수량"
        ws.cell(top2, c + 1).value = "계획금액"
        ws.cell(top2, c + 2).value = "실적수량"
        ws.cell(top2, c + 3).value = "실적금액"
        c += 4
    ws.cell(top2, c + 0).value = "계획수량"
    ws.cell(top2, c + 1).value = "계획금액"
    ws.cell(top2, c + 2).value = "실적수량"
    ws.cell(top2, c + 3).value = "실적금액"

    rr = top + 2
    for b in BRANDS:
        ws.cell(rr, 1).value = b
        c = 2
        for d in dates:
            v = prod_agg.get((b, d), {"plan_qty": 0.0, "plan_amt": 0.0, "act_qty": 0.0, "act_amt": 0.0})
            ws.cell(rr, c + 0).value = v["plan_qty"]
            ws.cell(rr, c + 1).value = v["plan_amt"]
            ws.cell(rr, c + 2).value = v["act_qty"]
            ws.cell(rr, c + 3).value = v["act_amt"]
            c += 4

        ws.cell(rr, c + 0).value = sum(prod_agg.get((b, d), {"plan_qty": 0.0})["plan_qty"] for d in dates)
        ws.cell(rr, c + 1).value = sum(prod_agg.get((b, d), {"plan_amt": 0.0})["plan_amt"] for d in dates)
        ws.cell(rr, c + 2).value = sum(prod_agg.get((b, d), {"act_qty": 0.0})["act_qty"] for d in dates)
        ws.cell(rr, c + 3).value = sum(prod_agg.get((b, d), {"act_amt": 0.0})["act_amt"] for d in dates)
        rr += 1

    ws.cell(rr, 1).value = "합계"
    c = 2
    for _ in dates:
        for k in range(4):
            col = get_column_letter(c + k)
            ws.cell(rr, c + k).value = f"=SUM({col}{top+2}:{col}{rr-1})"
        c += 4
    for k in range(4):
        col = get_column_letter(c + k)
        ws.cell(rr, c + k).value = f"=SUM({col}{top+2}:{col}{rr-1})"

    end_col = c + 3
    apply_table_style(ws, top, 1, rr, end_col, header_rows=2, total_row=rr)
    format_number_cells(ws, top + 2, 2, rr, end_col)
    pad_table_right(ws, top, rr, end_col, UNIFORM_END_COL)
    r = rr + 2

    chart_top = r
    ws.cell(chart_top, 1).value = "일자"
    ws.cell(chart_top, 2).value = "계획수량"
    ws.cell(chart_top, 3).value = "실적수량"
    ws.cell(chart_top, 4).value = "계획금액"
    ws.cell(chart_top, 5).value = "실적금액"
    rr = chart_top + 1
    for d in dates:
        plan_q = plan_a = act_q = act_a = 0.0
        for b in BRANDS:
            v = prod_agg.get((b, d), {"plan_qty": 0.0, "plan_amt": 0.0, "act_qty": 0.0, "act_amt": 0.0})
            plan_q += v["plan_qty"]
            plan_a += v["plan_amt"]
            act_q += v["act_qty"]
            act_a += v["act_amt"]
        ws.cell(rr, 1).value = d.strftime("%m/%d")
        ws.cell(rr, 2).value = plan_q
        ws.cell(rr, 3).value = act_q
        ws.cell(rr, 4).value = plan_a
        ws.cell(rr, 5).value = act_a
        rr += 1
    apply_table_style(ws, chart_top, 1, rr - 1, 5, header_rows=1)
    format_number_cells(ws, chart_top + 1, 2, rr - 1, 5)
    charts.append(("생산 상세 추세", chart_top, rr - chart_top))
    r = rr + 2

    return r, charts


def place_all_charts(ws, charts, start_row):
    row = start_row
    toggle = 0
    for title, table_top_row, n_rows in charts:
        anchor = f"{CHART_COL_LEFT}{row}" if toggle == 0 else f"{CHART_COL_RIGHT}{row}"
        chart = make_excel_style_combo_chart(
            ws,
            title=title,
            x_title="기간",
            y_left="수량",
            y_right="금액",
            table_top_row=table_top_row,
            table_left_col=1,
            n_rows=n_rows,
        )
        ws.add_chart(chart, anchor)

        if toggle == 0:
            toggle = 1
        else:
            toggle = 0
            row += CHART_ROW_GAP

    if toggle == 1:
        row += CHART_ROW_GAP

    return row


def build_excel_report(
    wb_src,
    month_start,
    month_end,
    week_start,
    week_end,
    day_start,
    day_end,
    prod_summary_start,
    prod_summary_end,
    prod_detail_start,
    prod_detail_end,
):
    orders_agg = collect_orders(wb_src)
    prod_agg = collect_production_plan_actual(wb_src)

    wb = Workbook()
    ws = wb.active
    ws.title = "생산일보"
    ws.sheet_view.showGridLines = False

    r = 1
    ws.cell(r, 1).value = "생산일보"
    set_cell_style(ws.cell(r, 1), font=TITLE_11, align=LEFT)
    r += 2

    r, order_charts = write_orders_section(ws, r, orders_agg, month_start, month_end, week_start, week_end, day_start, day_end)
    r, prod_charts = write_production_section(ws, r, prod_agg, prod_summary_start, prod_summary_end, prod_detail_start, prod_detail_end)

    charts_start_row = r + 2
    place_all_charts(ws, order_charts + prod_charts, charts_start_row)

    autosize_columns(ws, min_w=8, max_w=20)
    return wb, orders_agg, prod_agg


# =========================
# 웹 미리보기 데이터
# =========================
def build_preview_data(
    orders_agg,
    prod_agg,
    month_start,
    month_end,
    week_start,
    week_end,
    day_start,
    day_end,
    prod_summary_start,
    prod_summary_end,
    prod_detail_start,
    prod_detail_end,
):
    months = get_month_keys(month_start, month_end)
    rows1 = []
    for b in BRANDS:
        for mkr in MAKERS:
            row = {"브랜드-구분": f"{b}-{mkr}"}
            total_qty = total_amt = 0.0
            for mk in months:
                qty = amt = 0.0
                for (bb, mm, dt), vals in orders_agg.items():
                    if bb == b and mm == mkr and month_key(dt) == mk:
                        qty += vals["qty"]
                        amt += vals["amt"]
                row[f"{mk}_수량"] = qty
                row[f"{mk}_금액"] = amt
                total_qty += qty
                total_amt += amt
            row["합계_수량"] = total_qty
            row["합계_금액"] = total_amt
            rows1.append(row)
    df1 = pd.DataFrame(rows1)

    weeks = get_week_ranges(week_start, week_end)
    if len(weeks) > 5:
        weeks = weeks[-5:]
    rows2 = []
    for b in BRANDS:
        for mkr in MAKERS:
            row = {"브랜드-구분": f"{b}-{mkr}"}
            total_qty = total_amt = 0.0
            for ws_d, we_d in weeks:
                label = f"{ws_d.strftime('%m/%d')}~{we_d.strftime('%m/%d')}"
                qty = amt = 0.0
                for d in daterange(ws_d, we_d):
                    v = orders_agg.get((b, mkr, d), {"qty": 0.0, "amt": 0.0})
                    qty += v["qty"]
                    amt += v["amt"]
                row[f"{label}_수량"] = qty
                row[f"{label}_금액"] = amt
                total_qty += qty
                total_amt += amt
            row["합계_수량"] = total_qty
            row["합계_금액"] = total_amt
            rows2.append(row)
    df2 = pd.DataFrame(rows2)

    days = list(daterange(day_start, day_end))
    rows3 = []
    for b in BRANDS:
        for mkr in MAKERS:
            row = {"브랜드-구분": f"{b}-{mkr}"}
            total_qty = total_amt = 0.0
            for d in days:
                v = orders_agg.get((b, mkr, d), {"qty": 0.0, "amt": 0.0})
                label = d.strftime("%m/%d")
                row[f"{label}_수량"] = v["qty"]
                row[f"{label}_금액"] = v["amt"]
                total_qty += v["qty"]
                total_amt += v["amt"]
            row["합계_수량"] = total_qty
            row["합계_금액"] = total_amt
            rows3.append(row)
    df3 = pd.DataFrame(rows3)

    rows4 = []
    for b in BRANDS:
        plan_q = plan_a = act_q = act_a = 0.0
        for d in daterange(prod_summary_start, prod_summary_end):
            v = prod_agg.get((b, d), {"plan_qty": 0.0, "plan_amt": 0.0, "act_qty": 0.0, "act_amt": 0.0})
            plan_q += v["plan_qty"]
            plan_a += v["plan_amt"]
            act_q += v["act_qty"]
            act_a += v["act_amt"]
        rows4.append({
            "브랜드": b,
            "계획수량": plan_q,
            "계획금액": plan_a,
            "실적수량": act_q,
            "실적금액": act_a,
        })
    df4 = pd.DataFrame(rows4)

    prod_days = list(daterange(prod_detail_start, prod_detail_end))
    rows5 = []
    for b in BRANDS:
        row = {"브랜드": b}
        sum_pq = sum_pa = sum_aq = sum_aa = 0.0
        for d in prod_days:
            v = prod_agg.get((b, d), {"plan_qty": 0.0, "plan_amt": 0.0, "act_qty": 0.0, "act_amt": 0.0})
            label = d.strftime("%m/%d")
            row[f"{label}_계획수량"] = v["plan_qty"]
            row[f"{label}_계획금액"] = v["plan_amt"]
            row[f"{label}_실적수량"] = v["act_qty"]
            row[f"{label}_실적금액"] = v["act_amt"]
            sum_pq += v["plan_qty"]
            sum_pa += v["plan_amt"]
            sum_aq += v["act_qty"]
            sum_aa += v["act_amt"]
        row["합계_계획수량"] = sum_pq
        row["합계_계획금액"] = sum_pa
        row["합계_실적수량"] = sum_aq
        row["합계_실적금액"] = sum_aa
        rows5.append(row)
    df5 = pd.DataFrame(rows5)

    return df1, df2, df3, df4, df5


# =========================
# Streamlit UI
# =========================
st.set_page_config(page_title="생산일보 생성기", layout="wide")
st.title("생산일보 생성기")
st.markdown("엑셀 업로드 → 기간 설정 → 생산일보 생성 → 웹 표/그래프 확인 → 엑셀 다운로드")

uploaded_file = st.file_uploader("생산일보 사전 자료.xlsx 업로드", type=["xlsx"])

if uploaded_file is not None:
    st.info(f"업로드 파일: {uploaded_file.name}")

col1, col2 = st.columns(2)

with col1:
    st.subheader("수주 기간")
    month_start = st.date_input("1) 월별 수주현황 시작월", value=date(date.today().year, max(1, date.today().month - 2), 1))
    month_end = st.date_input("1) 월별 수주현황 종료월", value=date.today())

    week_start = st.date_input("2) 최근 5주간 주간합계 시작일", value=date.today() - timedelta(days=34))
    week_end = st.date_input("2) 최근 5주간 주간합계 종료일", value=date.today())

    day_start = st.date_input("3) 지난 1주일 수주 시작일", value=date.today() - timedelta(days=6))
    day_end = st.date_input("3) 지난 1주일 수주 종료일", value=date.today())

with col2:
    st.subheader("생산 기간")
    prod_summary_start = st.date_input("4) 생산 요약 시작일", value=date.today() - timedelta(days=7))
    prod_summary_end = st.date_input("4) 생산 요약 종료일", value=date.today())

    prod_detail_start = st.date_input("5) 생산 상세 시작일", value=date.today() - timedelta(days=6))
    prod_detail_end = st.date_input("5) 생산 상세 종료일", value=date.today())

st.subheader("수주 표 필터")
fcol1, fcol2 = st.columns(2)

with fcol1:
    selected_brands = st.multiselect("브랜드 선택", options=BRANDS, default=BRANDS)

with fcol2:
    selected_makers = st.multiselect("내외작 선택", options=MAKERS, default=MAKERS)

run_btn = st.button("생산일보 생성", type="primary")

if uploaded_file and run_btn:
    try:
        wb_src = load_workbook(filename=io.BytesIO(uploaded_file.read()), data_only=False)

        wb_result, orders_agg, prod_agg = build_excel_report(
            wb_src,
            month_start,
            month_end,
            week_start,
            week_end,
            day_start,
            day_end,
            prod_summary_start,
            prod_summary_end,
            prod_detail_start,
            prod_detail_end,
        )

        df1, df2, df3, df4, df5 = build_preview_data(
            orders_agg,
            prod_agg,
            month_start,
            month_end,
            week_start,
            week_end,
            day_start,
            day_end,
            prod_summary_start,
            prod_summary_end,
            prod_detail_start,
            prod_detail_end,
        )

        df_monthly_chart, df_weekly_chart, df_daily_chart, df_prod_summary_chart, df_prod_detail_chart = build_chart_data(
            orders_agg,
            prod_agg,
            month_start,
            month_end,
            week_start,
            week_end,
            day_start,
            day_end,
            prod_summary_start,
            prod_summary_end,
            prod_detail_start,
            prod_detail_end,
        )

        # 수주 표 필터 적용
        df1 = filter_order_rows_by_selection(df1, "브랜드-구분", selected_brands, selected_makers)
        df2 = filter_order_rows_by_selection(df2, "브랜드-구분", selected_brands, selected_makers)
        df3 = filter_order_rows_by_selection(df3, "브랜드-구분", selected_brands, selected_makers)

        # 수주 그래프 필터 적용
        if "내작" not in selected_makers:
            for col in ["내작수량", "내작금액"]:
                if col in df_monthly_chart.columns:
                    df_monthly_chart[col] = 0
                if col in df_weekly_chart.columns:
                    df_weekly_chart[col] = 0
                if col in df_daily_chart.columns:
                    df_daily_chart[col] = 0

        if "외주" not in selected_makers:
            for col in ["외주수량", "외주금액"]:
                if col in df_monthly_chart.columns:
                    df_monthly_chart[col] = 0
                if col in df_weekly_chart.columns:
                    df_weekly_chart[col] = 0
                if col in df_daily_chart.columns:
                    df_daily_chart[col] = 0

        # 하단 합계 추가
        df1 = add_total_row(df1, "브랜드-구분")
        df2 = add_total_row(df2, "브랜드-구분")
        df3 = add_total_row(df3, "브랜드-구분")
        df4 = add_total_row(df4, "브랜드")
        df5 = add_total_row(df5, "브랜드")

        st.success("생산일보 생성 완료")

        with st.expander("1) 월별 수주현황", expanded=True):
            st.dataframe(format_df_for_display(df1), use_container_width=True)
            show_combo_chart(
                df_monthly_chart,
                "월별 수주 추세",
                "기간",
                ["내작수량", "외주수량"],
                ["내작금액", "외주금액"],
            )

        with st.expander("2) 최근 5주간 주간 합계", expanded=True):
            st.dataframe(format_df_for_display(df2), use_container_width=True)
            show_combo_chart(
                df_weekly_chart,
                "최근 5주 주간 수주 추세",
                "기간",
                ["내작수량", "외주수량"],
                ["내작금액", "외주금액"],
            )

        with st.expander("3) 지난 1주일 수주 현황", expanded=True):
            st.dataframe(format_df_for_display(df3), use_container_width=True)
            show_combo_chart(
                df_daily_chart,
                "지난 1주일 수주 추세",
                "기간",
                ["내작수량", "외주수량"],
                ["내작금액", "외주금액"],
            )

        with st.expander("4) 생산 요약", expanded=True):
            st.dataframe(format_df_for_display(df4), use_container_width=True)
            show_combo_chart(
                df_prod_summary_chart,
                "생산 요약 계획 vs 실적",
                "브랜드",
                ["계획수량", "실적수량"],
                ["계획금액", "실적금액"],
            )

        with st.expander("5) 생산 상세", expanded=True):
            st.dataframe(format_df_for_display(df5), use_container_width=True)
            show_combo_chart(
                df_prod_detail_chart,
                "생산 상세 추세",
                "기간",
                ["계획수량", "실적수량"],
                ["계획금액", "실적금액"],
            )

        output = io.BytesIO()
        wb_result.save(output)
        output.seek(0)

        filename = f"생산일보_{date.today().strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label="엑셀 다운로드",
            data=output,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"오류 발생: {e}")
